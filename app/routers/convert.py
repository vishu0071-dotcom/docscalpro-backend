"""
File Conversion Router
Supports: PDF→DOCX, PDF→XLSX, DOCX→PDF, XLSX→PDF, Image→PDF
Uses: pdf2docx, pdfplumber, openpyxl, python-docx, reportlab, Pillow
"""

from fastapi import APIRouter, Depends, UploadFile, File, Form, HTTPException
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from typing import Optional
import os, uuid, shutil, tempfile
from datetime import datetime

from app.database import get_db, User, Conversion, ConversionStatus, PlanEnum
from app.routers.auth import verify_token

router = APIRouter()

UPLOAD_DIR = "/tmp/docscalpro/uploads"
OUTPUT_DIR = "/tmp/docscalpro/outputs"
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)

FREE_DAILY_LIMIT = 5
PREMIUM_DAILY_LIMIT = 999

# ─── Auth Helper ───────────────────────────────────────────────
def get_user_from_header(authorization: str, db: Session) -> User:
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Authorization required")
    user_id = verify_token(authorization.split(" ")[1])
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=401, detail="User not found")
    return user

def check_conversion_limit(user: User, db: Session):
    """Reset daily count if new day, then check limit"""
    today = datetime.utcnow().date()
    if user.last_conversion_date and user.last_conversion_date.date() != today:
        user.daily_conversions = 0

    limit = PREMIUM_DAILY_LIMIT if user.plan == PlanEnum.PREMIUM else FREE_DAILY_LIMIT
    if user.daily_conversions >= limit:
        raise HTTPException(
            status_code=403,
            detail=f"Daily limit reached ({limit} conversions/day for {user.plan.value} plan). Upgrade for unlimited conversions."
        )

# ─── Conversion Endpoint ───────────────────────────────────────
@router.post("")
async def convert_file(
    authorization: str = None,
    file: UploadFile = File(...),
    from_format: str = Form(...),
    to_format: str = Form(...),
    db: Session = Depends(get_db)
):
    user = get_user_from_header(authorization, db)
    check_conversion_limit(user, db)

    # Save uploaded file
    file_id = str(uuid.uuid4())
    input_ext = from_format.lower()
    if input_ext in ["jpg", "jpeg", "png"]:
        input_ext = "jpg"
    input_path = os.path.join(UPLOAD_DIR, f"{file_id}.{input_ext}")

    with open(input_path, "wb") as f:
        f.write(await file.read())

    # Track conversion in DB
    conversion = Conversion(
        id=file_id,
        user_id=user.id,
        from_format=from_format,
        to_format=to_format,
        original_filename=file.filename,
        file_size_kb=os.path.getsize(input_path) // 1024
    )
    db.add(conversion)
    db.flush()

    output_path = None
    try:
        output_path = await perform_conversion(input_path, from_format.upper(), to_format.upper(), file_id)
        output_filename = os.path.basename(output_path)

        # Update stats
        conversion.status = ConversionStatus.SUCCESS
        conversion.output_filename = output_filename
        user.total_conversions += 1
        user.daily_conversions += 1
        user.last_conversion_date = datetime.utcnow()
        db.commit()

        return {
            "success": True,
            "download_url": f"/api/v1/convert/download/{output_filename}",
            "file_name": output_filename,
            "error_message": None
        }

    except Exception as e:
        conversion.status = ConversionStatus.FAILED
        conversion.error_message = str(e)
        db.commit()
        raise HTTPException(status_code=500, detail=f"Conversion failed: {str(e)}")

    finally:
        # Cleanup input file
        if os.path.exists(input_path):
            os.remove(input_path)

async def perform_conversion(input_path: str, from_fmt: str, to_fmt: str, file_id: str) -> str:
    """Core conversion logic using open-source libraries"""
    output_path = None

    # ── PDF → DOCX ──────────────────────────────────────────────
    if from_fmt == "PDF" and to_fmt == "DOCX":
        from pdf2docx import Converter
        output_path = os.path.join(OUTPUT_DIR, f"{file_id}.docx")
        cv = Converter(input_path)
        cv.convert(output_path)
        cv.close()

    # ── PDF → XLSX ──────────────────────────────────────────────
    elif from_fmt == "PDF" and to_fmt == "XLSX":
        import pdfplumber
        import openpyxl
        output_path = os.path.join(OUTPUT_DIR, f"{file_id}.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Extracted Data"
        row_num = 1
        with pdfplumber.open(input_path) as pdf:
            for page in pdf.pages:
                # Extract tables
                tables = page.extract_tables()
                if tables:
                    for table in tables:
                        for table_row in table:
                            for col_idx, cell in enumerate(table_row):
                                ws.cell(row=row_num, column=col_idx + 1, value=cell or "")
                            row_num += 1
                        row_num += 1  # blank row between tables
                else:
                    # Extract text as fallback
                    text = page.extract_text()
                    if text:
                        for line in text.split('\n'):
                            ws.cell(row=row_num, column=1, value=line)
                            row_num += 1
        wb.save(output_path)

    # ── PDF → JPG/PNG ──────────────────────────────────────────
    elif from_fmt == "PDF" and to_fmt in ("JPG", "PNG"):
        import fitz  # PyMuPDF
        ext = to_fmt.lower()
        output_path = os.path.join(OUTPUT_DIR, f"{file_id}.{ext}")
        doc = fitz.open(input_path)
        page = doc[0]
        pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))  # 2x zoom for quality
        pix.save(output_path)
        doc.close()

    # ── DOCX → PDF ─────────────────────────────────────────────
    elif from_fmt == "DOCX" and to_fmt == "PDF":
        try:
            # Try LibreOffice if available (best quality)
            import subprocess
            result = subprocess.run(
                ["libreoffice", "--headless", "--convert-to", "pdf", "--outdir", OUTPUT_DIR, input_path],
                capture_output=True, timeout=30
            )
            # LibreOffice names output based on input filename
            base = os.path.splitext(os.path.basename(input_path))[0]
            lo_output = os.path.join(OUTPUT_DIR, f"{base}.pdf")
            output_path = os.path.join(OUTPUT_DIR, f"{file_id}.pdf")
            if os.path.exists(lo_output):
                os.rename(lo_output, output_path)
        except Exception:
            # Fallback: python-docx + reportlab (basic formatting)
            from docx import Document
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Paragraph
            from reportlab.lib.styles import getSampleStyleSheet
            doc = Document(input_path)
            output_path = os.path.join(OUTPUT_DIR, f"{file_id}.pdf")
            pdf_doc = SimpleDocTemplate(output_path, pagesize=A4)
            styles = getSampleStyleSheet()
            story = [Paragraph(para.text, styles['Normal']) for para in doc.paragraphs if para.text.strip()]
            pdf_doc.build(story)

    # ── XLSX → PDF ─────────────────────────────────────────────
    elif from_fmt == "XLSX" and to_fmt == "PDF":
        try:
            import subprocess
            result = subprocess.run(
                ["libreoffice", "--headless", "--convert-to", "pdf", "--outdir", OUTPUT_DIR, input_path],
                capture_output=True, timeout=30
            )
            base = os.path.splitext(os.path.basename(input_path))[0]
            lo_output = os.path.join(OUTPUT_DIR, f"{base}.pdf")
            output_path = os.path.join(OUTPUT_DIR, f"{file_id}.pdf")
            if os.path.exists(lo_output):
                os.rename(lo_output, output_path)
        except Exception:
            raise Exception("Excel to PDF conversion requires LibreOffice on server")

    # ── Image → PDF ─────────────────────────────────────────────
    elif from_fmt in ("JPG", "PNG", "JPEG") and to_fmt == "PDF":
        from PIL import Image
        output_path = os.path.join(OUTPUT_DIR, f"{file_id}.pdf")
        img = Image.open(input_path).convert("RGB")
        img.save(output_path, "PDF", resolution=100.0)

    else:
        raise Exception(f"Unsupported conversion: {from_fmt} → {to_fmt}")

    if not output_path or not os.path.exists(output_path):
        raise Exception("Conversion produced no output file")

    return output_path

# ─── Download Endpoint ─────────────────────────────────────────
@router.get("/download/{filename}")
async def download_file(filename: str, authorization: str = None, db: Session = Depends(get_db)):
    # Verify user is authenticated
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Authorization required")
    verify_token(authorization.split(" ")[1])

    # Security: prevent path traversal
    safe_filename = os.path.basename(filename)
    file_path = os.path.join(OUTPUT_DIR, safe_filename)

    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="File not found or expired")

    return FileResponse(
        path=file_path,
        filename=safe_filename,
        media_type="application/octet-stream"
    )
